"""Check the Excel dataset source for Healthy class."""
import openpyxl
from collections import Counter
import json

wb = openpyxl.load_workbook("dataset/DGA_Enterprise_Dataset.xlsx", data_only=True)
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    # Print header row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"  Headers: {headers}")
    
    # Find disease/class column
    class_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or '').lower()
        if any(k in h for k in ['disease', 'class', 'label', 'category', 'type', 'condition', 'health']):
            class_col = c
            print(f"  Class column found: column {c} = '{ws.cell(1, c).value}'")
            break
    
    if class_col:
        # Count unique values
        values = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, class_col).value
            if v is not None:
                values.append(str(v).strip())
        
        counter = Counter(values)
        print(f"  Unique class values ({len(counter)}):")
        for v, c in sorted(counter.items(), key=lambda x: -x[1]):
            print(f"    {v}: {c}")
        
        # Check for Healthy
        healthy_values = [v for v in values if 'healthy' in v.lower()]
        if healthy_values:
            print(f"\n  *** FOUND HEALTHY: {len(healthy_values)} occurrences ***")
        else:
            print(f"\n  No 'healthy' found in class column")
